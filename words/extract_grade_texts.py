from __future__ import annotations

from pathlib import Path
from typing import Iterable, List

from openpyxl import load_workbook


EXCLUDE_KEYWORDS = ("엑셀", "Sheet")


def iter_clean_rows(rows: Iterable[Iterable[object]]) -> List[str]:
    buffered_rows = list(rows)
    header_len = 0

    for row in buffered_rows:
        cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
        if cells:
            header_len = len(cells)
            break

    cleaned: List[str] = []
    for row in buffered_rows:
        values: List[str] = []
        taken = 0
        for cell in row:
            if header_len and taken >= header_len:
                break
            taken += 1
            if cell is None:
                continue

            text = str(cell).strip()
            if text:
                values.append(text)

        if not values:
            continue

        entry = " | ".join(values)
        if any(keyword in entry for keyword in EXCLUDE_KEYWORDS):
            continue

        cleaned.append(entry)
    return cleaned


def main() -> None:
    root = Path(__file__).resolve().parent
    workbook_path = root / "3~4학년 교과서 신출 어휘 및 의사소통 기능(이수지).xlsx"

    wb = load_workbook(workbook_path, data_only=True)
    worksheets = wb.worksheets

    data = {}
    grade_labels = [
        ("3학년", "신출 어휘", 0),
        ("3학년", "의사소통 기능", 1),
        ("4학년", "신출 어휘", 2),
        ("4학년", "의사소통 기능", 3),
    ]

    for grade, section, index in grade_labels:
        if index >= len(worksheets):
            entries: List[str] = []
        else:
            entries = iter_clean_rows(worksheets[index].iter_rows(values_only=True))

        data.setdefault(grade, {})[section] = entries

    for grade, sections in data.items():
        output_path = root / f"{grade}_신출어휘_의사소통기능.txt"
        with output_path.open("w", encoding="utf-8", newline="\n") as f:
            for title, rows in sections.items():
                f.write(f"{title} (총 {len(rows)}항목)\n")
                for entry in rows:
                    f.write(entry + "\n")
                f.write("\n")


if __name__ == "__main__":
    main()

